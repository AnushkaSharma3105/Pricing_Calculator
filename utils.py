import base64
import pandas as pd
import io
from datetime import datetime
from pathlib import Path

LOGO_PATH = Path(__file__).parent / "assets" / "ttbs_logo.png"
LOGO_TOP_ROWS = 3  # empty rows inserted above the unchanged quotation template
LOGO_DISPLAY_HEIGHT_PX = 42  # embedded logo height; width keeps aspect ratio

# Default column widths (chars) used in BOQ exports
BOQ_COL_WIDTHS = [22, 18, 20, 22, 10, 12, 14, 20, 8, 28, 22]


def get_logo_base64():
    """Return base64-encoded TTBS logo for inline HTML display."""
    if LOGO_PATH.exists():
        return base64.b64encode(LOGO_PATH.read_bytes()).decode()
    return ""


def format_inr(amount):
    """Format number as Indian Rupees"""
    return f"₹ {amount:,.2f}"


def generate_quotation_id():
    """Generate a unique quotation ID based on timestamp"""
    now = datetime.now()
    return f"QUO-{now.strftime('%Y%m%d-%H%M%S')}"


def build_summary_dataframe(product, flavour, specs, config, breakdown):
    """Build a clean summary dataframe for display and export"""

    rows = []

    # Product info
    rows.append({"Component": "Product", "Details": product, "Amount (INR)": ""})
    rows.append({"Component": "Flavour", "Details": flavour, "Amount (INR)": ""})

    # Specs
    for k, v in specs.items():
        rows.append({"Component": k, "Details": str(v), "Amount (INR)": ""})

    # Config
    for k, v in config.items():
        rows.append({"Component": k, "Details": str(v), "Amount (INR)": ""})

    # Blank separator
    rows.append({"Component": "", "Details": "", "Amount (INR)": ""})

    # Pricing breakdown
    rows.append({"Component": "--- PRICE BREAKDOWN ---", "Details": "", "Amount (INR)": ""})
    for k, v in breakdown.items():
        if k not in ("Quantity", "Grand Total"):
            rows.append({
                "Component": k,
                "Details": "",
                "Amount (INR)": f"{v:,.2f}" if isinstance(v, (int, float)) else v
            })

    rows.append({"Component": "", "Details": "", "Amount (INR)": ""})
    rows.append({
        "Component": "Quantity",
        "Details": str(breakdown.get("Quantity", 1)),
        "Amount (INR)": ""
    })
    rows.append({
        "Component": "GRAND TOTAL",
        "Details": "",
        "Amount (INR)": f"{breakdown.get('Grand Total', 0):,.2f}"
    })

    return pd.DataFrame(rows)


def export_to_csv(df):
    """Export dataframe to CSV bytes"""
    return df.to_csv(index=False).encode("utf-8")


def _prepare_logo_image_bytes():
    """Resize the shared TTBS logo for crisp embedding (same file as the website)."""
    if not LOGO_PATH.exists():
        return None, 0, 0

    try:
        from PIL import Image
    except ImportError:
        return None, 0, 0

    with Image.open(LOGO_PATH) as src:
        aspect = src.width / src.height
        display_h = LOGO_DISPLAY_HEIGHT_PX
        display_w = max(1, int(display_h * aspect))
        resized = src.convert("RGBA").resize(
            (display_w, display_h), Image.Resampling.LANCZOS,
        )
        buf = io.BytesIO()
        resized.save(buf, format="PNG", dpi=(96, 96))
        buf.seek(0)
        return buf, display_w, display_h


def _embed_logo_at_top(xlsx_bytes, sheet_name):
    """Insert embedded TTBS logo in reserved top rows; template below is untouched."""
    logo_buf, display_w, display_h = _prepare_logo_image_bytes()
    if not logo_buf:
        return xlsx_bytes

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        if sheet_name not in wb.sheetnames:
            return xlsx_bytes

        ws = wb[sheet_name]
        ws._images.clear()

        total_pts = (display_h + 8) * 72 / 96
        row_pts = total_pts / LOGO_TOP_ROWS
        for r in range(LOGO_TOP_ROWS):
            ws.row_dimensions[r + 1].height = row_pts

        y_pad = max(2, int((total_pts * 96 / 72 - display_h) / 2))
        img = XLImage(logo_buf)
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=0,
                colOff=pixels_to_EMU(8),
                row=0,
                rowOff=pixels_to_EMU(y_pad),
            ),
            ext=XDRPositiveSize2D(
                pixels_to_EMU(display_w),
                pixels_to_EMU(display_h),
            ),
        )
        ws.add_image(img)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()
    except Exception:
        return xlsx_bytes


def _reserve_logo_rows(worksheet):
    """Leave blank rows at the top so the logo never overlaps template content."""
    for r in range(LOGO_TOP_ROWS):
        worksheet.set_row(r, 16)


def export_to_excel(df, product, flavour, quotation_id):
    """Export dataframe to Excel bytes with formatting"""
    output = io.BytesIO()
    title_rows = 4
    data_startrow = LOGO_TOP_ROWS + title_rows

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(
            writer, index=False, sheet_name="Quotation",
            startrow=data_startrow,
        )

        workbook = writer.book
        worksheet = writer.sheets["Quotation"]

        title_fmt = workbook.add_format({
            "bold": True, "font_size": 16,
            "font_color": "#FFFFFF", "bg_color": "#1B3A6B",
            "align": "center", "valign": "vcenter"
        })
        subtitle_fmt = workbook.add_format({
            "bold": True, "font_size": 11,
            "font_color": "#1B3A6B"
        })
        header_fmt = workbook.add_format({
            "bold": True, "bg_color": "#1B3A6B",
            "font_color": "#FFFFFF", "border": 1,
            "align": "center"
        })
        total_fmt = workbook.add_format({
            "bold": True, "bg_color": "#FFF3CD",
            "font_color": "#856404", "border": 1,
            "num_format": "#,##0.00"
        })

        _reserve_logo_rows(worksheet)

        worksheet.merge_range(LOGO_TOP_ROWS, 0, LOGO_TOP_ROWS, 2, "PRICE QUOTATION", title_fmt)
        worksheet.write(LOGO_TOP_ROWS + 1, 0, f"Quotation ID: {quotation_id}", subtitle_fmt)
        worksheet.write(LOGO_TOP_ROWS + 2, 0, f"Product: {product}  |  Flavour: {flavour}", subtitle_fmt)
        worksheet.write(
            LOGO_TOP_ROWS + 3, 0,
            f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}",
            subtitle_fmt,
        )

        for col_num, col_name in enumerate(df.columns):
            worksheet.write(data_startrow, col_num, col_name, header_fmt)

        worksheet.set_column("A:A", 35)
        worksheet.set_column("B:B", 25)
        worksheet.set_column("C:C", 20)

        for row_idx, row in df.iterrows():
            if str(row.get("Component", "")).strip() == "GRAND TOTAL":
                excel_row = row_idx + data_startrow + 1
                worksheet.write(excel_row, 0, "GRAND TOTAL", total_fmt)
                worksheet.write(excel_row, 1, "", total_fmt)
                worksheet.write(excel_row, 2, row["Amount (INR)"], total_fmt)

    return _embed_logo_at_top(output.getvalue(), "Quotation")


def build_quote_export_dataframe(items):
    rows = []
    for index, item in enumerate(items, start=1):
        rows.append({
            "S.No.": index,
            "Category": item.get("Category", ""),
            "Product": item.get("Product", ""),
            "Flavour": item.get("Flavour", ""),
            "Element": item.get("Element", ""),
            "Hypervisor": item.get("Hypervisor", ""),
            "Operating System": item.get("Operating System", ""),
            "Pricing Tier": item.get("Pricing Tier", ""),
            "Storage Type": item.get("Storage Type", ""),
            "Storage (GB)": item.get("Storage (GB)", 0),
            "Backup Type": item.get("Backup Type", ""),
            "Backup (GB)": item.get("Backup (GB)", 0),
            "Firewall Type": item.get("Firewall Type", ""),
            "Firewall Bandwidth (Mbps)": item.get("Firewall Bandwidth (Mbps)", 0),
            "Firewall Cost (INR)": item.get("Firewall Cost (INR)", 0.0),
            "Public IPs": item.get("Public IPs", 0),
            "Quantity": item.get("Quantity", 1),
            "vCPU": item.get("vCPU", ""),
            "RAM (GB)": item.get("RAM (GB)", ""),
            # ── Step 3: Network & Security ──
            "Network Element": item.get("Network Element", ""),
            "Network Feature": item.get("Network Feature", ""),
            "Network Sub Type": item.get("Network Sub Type", ""),
            "Bandwidth (Mbps)": item.get("Bandwidth (Mbps)", 0),
            "Network Cost (INR)": item.get("Network Cost (INR)", 0.0),
            # ── Step 3: Licenses ──
            "License Element": item.get("License Element", ""),
            "License Sub Type": item.get("License Sub Type", ""),
            "License Qty": item.get("License Qty", 0),
            "License Cost (INR)": item.get("License Cost (INR)", 0.0),
            # ── Step 3: Backup Storage ──
            "Backup Storage Model": item.get("Backup Storage Model", ""),
            "Backup Storage (GB)": item.get("Backup Storage (GB)", 0),
            "Backup Storage Cost (INR)": item.get("Backup Storage Cost (INR)", 0.0),
            # ── Step 3: Network Elements ──
            "Network Element Type": item.get("Network Element Type", ""),
            "Network Element Cost (INR)": item.get("Network Element Cost (INR)", 0.0),
            # ── Step 3: Management ──
            "Management Type": item.get("Management Type", ""),
            "Management Qty": item.get("Management Qty", 0),
            "Management Cost (INR)": item.get("Management Cost (INR)", 0.0),
            # ── Step 3: Miscellaneous ──
            "Misc Element": item.get("Misc Element", ""),
            "Misc Qty": item.get("Misc Qty", 0),
            "Misc Cost (INR)": item.get("Misc Cost (INR)", 0.0),
            # ── Totals ──
            "Additional Services Total (INR)": item.get("Additional Services Total (INR)", 0.0),
            "Line Total (INR)": item.get("Line Total (INR)", 0.0),
        })

    df = pd.DataFrame(rows)
    return df


def export_quote_to_csv(df, grand_total=None):
    """Export in BOQ format as CSV"""
    lines = []

    lines.append("PRICE QUOTATION - AETHERPRICE")
    lines.append("")
    lines.append("Account Manager,,BD,,Solution Architect,,")
    lines.append("Service Model,IPC,Contract,,IDC Location,,")
    lines.append("")

    # Network & Security
    lines.append("Network and Security Services,,,,,,,,,,MRC/1Y Price")
    lines.append("Element,Feature,Sub Type,Make,Model,Description,Unit,Qty,Remark,Cost (INR)")
    net_rows = df[df["Network Element"].notna() & (df["Network Element"] != "")]
    if not net_rows.empty:
        for _, row in net_rows.iterrows():
            lines.append(
                f"{row.get('Network Element','')},{row.get('Network Feature','')},{row.get('Network Sub Type','')},"
                f",,{row.get('Network Sub Type','')},Mbps,{row.get('Bandwidth (Mbps)',0)},,"
                f"{row.get('Network Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,,,,,")
    lines.append("")

    # Firewall
    lines.append("Firewall,,,,,,,,,,")
    lines.append("Type,Bandwidth (Mbps),Remark,,,,,,,,Cost (INR)")
    fw_rows = df[df["Firewall Type"].notna() & (df["Firewall Type"] != "")]
    if not fw_rows.empty:
        for _, row in fw_rows.iterrows():
            lines.append(
                f"{row.get('Firewall Type','')},{row.get('Firewall Bandwidth (Mbps)',0)},,,,,,,,,"
                f"{row.get('Firewall Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,,,,,")
    lines.append("")

    # VM Section
    lines.append("Vayu Private Cloud")
    lines.append("Element,Hypervisor,Environment,Operating System,vCPU,Memory (GB),Root Disk,Additional Storage (GB),Qty,Remark,Total VM + Disk Price/Month")
    vm_rows = df[df["Flavour"].notna() & (df["Flavour"] != "")]
    for _, row in vm_rows.iterrows():
        lines.append(
            f"{row.get('Element','ICS')},{row.get('Hypervisor','Open Stack')},Prod,"
            f"{row.get('Operating System','')},{row.get('vCPU','')},{row.get('RAM (GB)','')},"
            f",{row.get('Storage (GB)',0)},{row.get('Quantity',1)},{row.get('Storage Type','')},"
            f"{row.get('Line Total (INR)',0)}"
        )
    lines.append("")

    # Software & Licenses
    lines.append("Software and Licenses,,,,,,,,,,")
    lines.append("Element,Sub Type,Description,Unit,Qty,Remark,Cost (INR)")
    lic_rows = df[df["License Element"].notna() & (df["License Element"] != "")]
    if not lic_rows.empty:
        for _, row in lic_rows.iterrows():
            lines.append(
                f"{row.get('License Element','')},{row.get('License Sub Type','')},,"
                f",{row.get('License Qty',0)},,{row.get('License Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,,,")
    lines.append("")

    # Backup Storage
    lines.append("Backup Storage ZRS")
    lines.append("Element,Make,Model,Storage Configuration,Description,Unit,Qty,Remark,Cost (INR)")
    bk_rows = df[df["Backup Storage Model"].notna() & (df["Backup Storage Model"] != "")]
    if not bk_rows.empty:
        for _, row in bk_rows.iterrows():
            lines.append(
                f"{row.get('Element','ICS')},BET,{row.get('Backup Storage Model','')},"
                f"Object-Resilient,,GB,{row.get('Backup Storage (GB)',0)},,"
                f"{row.get('Backup Storage Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,,,,")
    lines.append("")

    # Network Elements
    lines.append("Network Element")
    lines.append("Element,Description,Unit,Qty,Remark,Cost (INR)")
    ne_rows = df[df["Network Element Type"].notna() & (df["Network Element Type"] != "")]
    if not ne_rows.empty:
        for _, row in ne_rows.iterrows():
            lines.append(
                f"{row.get('Network Element Type','')},,,1,,"
                f"{row.get('Network Element Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,")
    lines.append("")

    # Management
    lines.append("Management Services")
    lines.append("Element,Description,Unit,Qty,Remark,Cost (INR)")
    mg_rows = df[df["Management Type"].notna() & (df["Management Type"] != "")]
    if not mg_rows.empty:
        for _, row in mg_rows.iterrows():
            lines.append(
                f"{row.get('Management Type','')},,,{row.get('Management Qty',0)},,"
                f"{row.get('Management Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,")
    lines.append("")

    # Miscellaneous
    lines.append("Miscellaneous")
    lines.append("Element,Description,Unit,Qty,Remark,Cost (INR)")
    mi_rows = df[df["Misc Element"].notna() & (df["Misc Element"] != "")]
    if not mi_rows.empty:
        for _, row in mi_rows.iterrows():
            lines.append(
                f"{row.get('Misc Element','')},,,{row.get('Misc Qty',0)},,"
                f"{row.get('Misc Cost (INR)',0)}"
            )
    else:
        lines.append(",,,,,,")
    lines.append("")

    # Grand Total
    if grand_total is not None:
        lines.append(f"GRAND TOTAL,,,,,,,,,{grand_total}")

    return "\n".join(lines).encode("utf-8")


def export_quote_to_excel(df, quotation_id, grand_total):
    """Export in BOQ format as Excel matching OUTPUT_FORMAT.xlsx structure"""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("BOQ")

        # Formats
        title_fmt = workbook.add_format({
            "bold": True, "font_size": 16,
            "font_color": "#FFFFFF", "bg_color": "#1B3A6B",
            "align": "center", "valign": "vcenter"
        })
        subtitle_fmt = workbook.add_format({
            "bold": True, "font_size": 11, "font_color": "#1B3A6B"
        })
        header_fmt = workbook.add_format({
            "bold": True, "bg_color": "#1B3A6B",
            "font_color": "#FFFFFF", "border": 1,
            "align": "center", "text_wrap": True
        })
        section_fmt = workbook.add_format({
            "bold": True, "font_size": 11,
            "bg_color": "#BDD7EE", "font_color": "#1B3A6B",
            "border": 1
        })
        meta_label_fmt = workbook.add_format({
            "bold": True, "font_color": "#1B3A6B"
        })
        meta_value_fmt = workbook.add_format({
            "font_color": "#1B3A6B"
        })
        data_fmt = workbook.add_format({
            "border": 1, "text_wrap": True
        })
        number_fmt = workbook.add_format({
            "border": 1, "num_format": "#,##0.00"
        })
        sno_fmt = workbook.add_format({
        "border": 1, "num_format": "0"
        })
        total_fmt = workbook.add_format({
            "bold": True, "bg_color": "#FFF3CD",
            "font_color": "#856404", "border": 1,
            "num_format": "#,##0.00"
        })
        total_label_fmt = workbook.add_format({
            "bold": True, "bg_color": "#FFF3CD",
            "font_color": "#856404", "border": 1
        })

        boq_col_widths = [8, 22, 18, 20, 22, 10, 12, 14, 20, 8, 28, 22]
        worksheet.set_column("A:A", boq_col_widths[0])
        worksheet.set_column("B:B", boq_col_widths[1])
        worksheet.set_column("C:C", boq_col_widths[2])
        worksheet.set_column("D:D", boq_col_widths[3])
        worksheet.set_column("E:E", boq_col_widths[4])
        worksheet.set_column("F:F", boq_col_widths[5])
        worksheet.set_column("G:G", boq_col_widths[6])
        worksheet.set_column("H:H", boq_col_widths[7])
        worksheet.set_column("I:I", boq_col_widths[8])
        worksheet.set_column("J:J", boq_col_widths[9])
        worksheet.set_column("K:K", boq_col_widths[10])
        worksheet.set_column("L:L", boq_col_widths[11])

        _reserve_logo_rows(worksheet)

        row = LOGO_TOP_ROWS
        worksheet.merge_range(row, 0, row, 11, "PRICE QUOTATION", title_fmt)
        row += 1
        worksheet.write(row, 0, f"Quotation ID: {quotation_id}", subtitle_fmt)
        row += 1
        worksheet.write(row, 0, f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", subtitle_fmt)
        row += 1
        worksheet.write(row, 0, "Account Manager", meta_label_fmt)
        worksheet.write(row, 2, "BD", meta_label_fmt)
        worksheet.write(row, 4, "Solution Architect", meta_label_fmt)
        row += 1
        worksheet.write(row, 0, "Service Model", meta_label_fmt)
        worksheet.write(row, 1, "IPC", meta_value_fmt)
        worksheet.write(row, 2, "Contract", meta_label_fmt)
        worksheet.write(row, 4, "IDC Location", meta_label_fmt)
        row += 1

        def write_section_header(ws, r, title):
            ws.merge_range(r, 0, r, 11, title, section_fmt)
            return r + 1

        def write_col_headers(ws, r, headers):
            for c, h in enumerate(headers):
                ws.write(r, c, h, header_fmt)
            return r + 1

        def write_data_row(ws, r, values, fmts=None):
            for c, v in enumerate(values):
                fmt = fmts[c] if fmts and c < len(fmts) else (number_fmt if isinstance(v, (int, float)) else data_fmt)
                ws.write(r, c, v, fmt)
            return r + 1
        
        def row_fmts(values):
            fmts = [sno_fmt]
            for v in values[1:]:
                fmts.append(number_fmt if isinstance(v, (int, float)) else data_fmt)
            return fmts

        
        # SECTION 1: Network & Security Services
        
        row = write_section_header(worksheet, row, "Network and Security Services")
        ns_headers = ["Element", "Feature", "Sub Type", "Make", "Model", None,
                      "Description", "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, ns_headers)
        net_rows = df[df["Network Element"].notna() & (df["Network Element"] != "")]
        if not net_rows.empty:
            for _, r2 in net_rows.iterrows():
                vals = [
                    r2.get("Network Element", ""),
                    r2.get("Network Feature", ""),
                    r2.get("Network Sub Type", ""),
                    "", "", "",
                    r2.get("Network Sub Type", ""),
                    r2.get("Bandwidth (Mbps)", 0),
                    "Mbps",
                    "",
                    r2.get("Network Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals)
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 1B: Firewall
        
        row = write_section_header(worksheet, row, "Firewall")
        fw_headers = ["Type", "Bandwidth (Mbps)", None, None, None, None, None,
                      None, None, "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, fw_headers)
        fw_rows = df[df["Firewall Type"].notna() & (df["Firewall Type"] != "")]
        if not fw_rows.empty:
            for _, r2 in fw_rows.iterrows():
                vals = [
                    r2.get("Firewall Type", ""),
                    r2.get("Firewall Bandwidth (Mbps)", 0),
                    "", "", "", "", "", "", "",
                    "",
                    r2.get("Firewall Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals)
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 2: Vayu Private Cloud (VM)
        
        row = write_section_header(worksheet, row, "Vayu Private Cloud")
        vm_headers = ["Element", "Hypervisor", "Environment", "Operating System",
                      "vCPU", "Memory (GB)", "Root Disk", "Additional Storage (GB)",
                      "Qty", "Remark", "Total VM + Disk Price/Month"]
        row = write_col_headers(worksheet, row, vm_headers)
        vm_rows = df[df["Flavour"].notna() & (df["Flavour"] != "")]
        for _, r2 in vm_rows.iterrows():
            vals = [
                r2.get("Element", "ICS"),
                r2.get("Hypervisor", "Open Stack"),
                "Prod",
                r2.get("Operating System", ""),
                r2.get("vCPU", ""),
                r2.get("RAM (GB)", ""),
                "",
                r2.get("Storage (GB)", 0),
                r2.get("Quantity", 1),
                r2.get("Storage Type", ""),
                r2.get("Line Total (INR)", 0),
            ]
            row = write_data_row(worksheet, row, vals)
        row += 1

        
        # SECTION 3: Software & Licenses
        
        row = write_section_header(worksheet, row, "Software and Licenses")
        lic_headers = ["Element", "Sub Type", "Description", None, None, None, None,
                       "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, lic_headers)
        lic_rows = df[df["License Element"].notna() & (df["License Element"] != "")]
        if not lic_rows.empty:
            for _, r2 in lic_rows.iterrows():
                vals = [
                    r2.get("License Element", ""),
                    r2.get("License Sub Type", ""),
                    "",
                    "", "", "", "",
                    "# of Licenses",
                    r2.get("License Qty", 0),
                    "",
                    r2.get("License Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals)
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 4: Backup Storage
        
        row = write_section_header(worksheet, row, "Backup Storage")
        bk_headers = ["Element", "Make", "Model", "Storage Configuration",
                      "Description", None, None, "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, bk_headers)
        bk_rows = df[df["Backup Storage Model"].notna() & (df["Backup Storage Model"] != "")]
        if not bk_rows.empty:
            for _, r2 in bk_rows.iterrows():
                vals = [
                    r2.get("Element", "ICS"),
                    "BET",
                    r2.get("Backup Storage Model", ""),
                    "Object-Resilient",
                    "",
                    "", "",
                    "GB",
                    r2.get("Backup Storage (GB)", 0),
                    "",
                    r2.get("Backup Storage Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals, fmts=row_fmts(vals))
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 5: Network Elements
        
        row = write_section_header(worksheet, row, "Network Element")
        ne_headers = ["Element", "Description", None, None, None, None, None,
                      "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, ne_headers)
        ne_rows = df[df["Network Element Type"].notna() & (df["Network Element Type"] != "")]
        if not ne_rows.empty:
            for _, r2 in ne_rows.iterrows():
                vals = [
                    r2.get("Network Element Type", ""),
                    "", "", "", "", "", "",
                    "", 1, "",
                    r2.get("Network Element Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals, fmts=row_fmts(vals))
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 6: Management Services
        
        row = write_section_header(worksheet, row, "Management Services")
        mg_headers = ["Element", "Description", None, None, None, None, None,
                      "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, mg_headers)
        mg_rows = df[df["Management Type"].notna() & (df["Management Type"] != "")]
        if not mg_rows.empty:
            for _, r2 in mg_rows.iterrows():
                vals = [
                    r2.get("Management Type", ""),
                    "", "", "", "", "", "",
                    "",
                    r2.get("Management Qty", 0),
                    "",
                    r2.get("Management Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals, fmts=row_fmts(vals))
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # SECTION 7: Miscellaneous
        
        row = write_section_header(worksheet, row, "Miscellaneous Items")
        mi_headers = ["Element", "Description", None, None, None, None, None,
                      "Unit", "Qty", "Remark", "Cost (INR)"]
        row = write_col_headers(worksheet, row, mi_headers)
        mi_rows = df[df["Misc Element"].notna() & (df["Misc Element"] != "")]
        if not mi_rows.empty:
            for _, r2 in mi_rows.iterrows():
                vals = [
                    r2.get("Misc Element", ""),
                    "", "", "", "", "", "",
                    "",
                    r2.get("Misc Qty", 0),
                    "",
                    r2.get("Misc Cost (INR)", 0),
                ]
                row = write_data_row(worksheet, row, vals, fmts=row_fmts(vals))
        else:
            row = write_data_row(worksheet, row, [""] * 11)
        row += 1

        
        # GRAND TOTAL ROW
        
        worksheet.merge_range(row, 0, row, 10, "GRAND TOTAL", total_label_fmt)
        worksheet.write(row, 11, grand_total, total_fmt)

    return _embed_logo_at_top(output.getvalue(), "BOQ")